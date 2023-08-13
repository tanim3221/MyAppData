import os
from datetime import datetime

import openpyxl
import pandas as pd
import PyPDF2
from colorama import Back, Style
from openpyxl import load_workbook
from PIL import Image
from PyPDF2 import PdfMerger
from reportlab.lib.pagesizes import A4, letter
from reportlab.pdfgen import canvas


def initial():
    intinput = input(
        '''\nWhat do you want:\n'''
        '''1. Merge excel files\n'''
        '''2. Clean excel files\n'''
        '''3. Clean and Merge excel files\n'''
        '''4. Filter with keyword and Merge excel files\n'''
        '''5. Merge excel sheets\n'''
        '''6. Multiple PDF Merge\n'''
        '''7. Split PDF Defined pages\n'''
        '''8. Split PDF all pages\n'''
        '''9. Make JPG to PDF\n'''
        '''\nEnter your option (1/2/3/4/5/6/7/8/9): '''
        )

    if intinput=="1":
        intext = "\nMerge Excel files\n"
        intext = Back.GREEN + intext + Style.RESET_ALL
        print(intext)
        mergeExcel()
    elif intinput =="2":
        intext = "\nClean Excel files\n"
        intext = Back.GREEN + intext + Style.RESET_ALL
        print(intext)
        cleanExcel()
    elif intinput =="3":
        intext = "\nClean and Merge Excel files"
        intext = Back.GREEN + intext + Style.RESET_ALL
        print(intext)
        clean_and_merge()
    elif intinput =="4":
        intext = "\nFilter and Merge Excel files\n"
        intext = Back.GREEN + intext + Style.RESET_ALL
        print(intext)
        filter_and_merge_key()
    elif intinput =="5":
        intext = "\nMerge Excel Sheets\n"
        intext = Back.GREEN + intext + Style.RESET_ALL
        print(intext)
        merge_sheets()
    elif intinput =="6":
        intext = "\nMultiple PDF Merge\n"
        intext = Back.GREEN + intext + Style.RESET_ALL
        print(intext)
        mergePDF()
    elif intinput =="7":
        intext = "\nSplit PDF Defined pages\n"
        intext = Back.GREEN + intext + Style.RESET_ALL
        print(intext)
        pdfSplitDefined()
    elif intinput =="8":
        intext = "\nSplit PDF all pages\n"
        intext = Back.GREEN + intext + Style.RESET_ALL
        print(intext)
        pdfSplitAll()
    elif intinput =="9":
        intext = "\nMake JPG to PDF\n"
        intext = Back.GREEN + intext + Style.RESET_ALL
        print(intext)
        jpgtopdf()
    elif not intinput.isdigit() or not (1 <= int(intinput) <= 8):
        if not intinput.isdigit():
            intext = "\nSorry! You have to select any option here. Executing again..."
        else:
            intext = "\nTo select any option, you have to enter a number between 1 and 8. Executing again..."
        intext = Back.RED + intext + Style.RESET_ALL
        print(intext)
        initial()


def jpgtopdf():
    curr_time = datetime.now().strftime("%Y%m%d%H%M%S")

    # Input folder containing image files (JPG, PNG, etc.)
    input_folder = input("\nEnter image folder directory: ")

    # Get a list of image file paths from the input folder
    image_paths = [os.path.join(input_folder, file) for file in os.listdir(input_folder) if file.lower().endswith(('.jpg', '.png', '.jpeg', '.gif', '.bmp'))]

    # Output PDF file name based on the input folder name
    pdf_output = os.path.join(input_folder, f"{os.path.basename(input_folder)}_"+curr_time+".pdf")

    # Create a PDF using ReportLab
    c = canvas.Canvas(pdf_output, pagesize=letter)

    for image_path in image_paths:
        image = Image.open(image_path)
        image_width, image_height = image.size

        if image_width < A4[0] and image_height < A4[1]:
            page_size = A4  # Set to A4 if smaller than A4
        else:
            page_size = (image_width, image_height)

        c.setPageSize(page_size)
        c.drawImage(image_path, 0, 0, width=image_width, height=image_height, preserveAspectRatio=True, anchor='nw')
        c.showPage()

    c.save()

    print(f"\nPDF created and save to : {pdf_output}")


def pdfSplitAll():
    curr_time = datetime.now().strftime("%Y%m%d%H%M%S")
    
    def split_pdf(input_path, output_folder, pages):
        pdf = PyPDF2.PdfReader(input_path)
        
        for page_num in range(len(pdf.pages)):
            output_pdf_name = f"Page_{page_num + 1}_{curr_time}.pdf"
            output_pdf_path = os.path.join(output_folder, output_pdf_name)
            
            with open(output_pdf_path, "wb") as output_file:
                writer = PyPDF2.PdfWriter()
                writer.add_page(pdf.pages[page_num])
                writer.write(output_file)

    folder_path = input("\nEnter folder path containing the PDF files:\n")
    pdf_files = [pdf_file for pdf_file in os.listdir(folder_path) if pdf_file.endswith('.pdf')]
    pdf_files.sort()

    if not pdf_files:
        print("No PDF files found in the folder.")
        return

    print("\nPDF files found in the folder:\n")
    for index, pdf_file in enumerate(pdf_files, start=1):
        print(f"{index}. {pdf_file}")

    pdf_choice = input("\nEnter the number of the PDF to split or type 'cancel': ")
    if pdf_choice.lower() == "no":
        print("\nPDF splitting canceled. Thank you!")
        return

    try:
        pdf_index = int(pdf_choice) - 1
        if pdf_index < 0 or pdf_index >= len(pdf_files):
            print("Invalid PDF number.")
            return

        selected_pdf = pdf_files[pdf_index]
        input_pdf_path = os.path.join(folder_path, selected_pdf)
        output_folder = os.path.join(folder_path, f"{os.path.splitext(selected_pdf)[0]}_{curr_time}")
        os.makedirs(output_folder, exist_ok=True)
        
        pages_to_split = input("\nDo you want to split these PDFs? (yes/no): ")

        if pages_to_split.lower() == "yes":
            split_pdf(input_pdf_path, output_folder, pages_to_split)
            print(f"\nPDF split into individual pages and saved to {output_folder}")

    except ValueError:
        print("Invalid input.")

def pdfSplitDefined():
    curr_time = datetime.now().strftime("%Y%m%d%H%M%S")
    
    def split_pdf(input_path, output_path, pages):
        pdf = PyPDF2.PdfReader(input_path)
        
        with open(output_path, "wb") as output_file:
            writer = PyPDF2.PdfWriter()
            
            for page_range in pages.split(','):
                if '-' in page_range:
                    start, end = map(int, page_range.split('-'))
                    for page_num in range(start - 1, end):
                        writer.add_page(pdf.pages[page_num])
                else:
                    page_num = int(page_range) - 1
                    writer.add_page(pdf.pages[page_num])
            
            writer.write(output_file)

    folder_path = input("\nEnter folder path containing the PDF files:\n")
    pdf_files = [pdf_file for pdf_file in os.listdir(folder_path) if pdf_file.endswith('.pdf')]
    pdf_files.sort()

    if not pdf_files:
        print("No PDF files found in the folder.")
        return

    print("\nPDF files found in the folder:\n")
    for index, pdf_file in enumerate(pdf_files, start=1):
        print(f"{index}. {pdf_file}")

    pdf_choice = input("\nEnter the number of the PDF to split or type 'cancel': ")
    if pdf_choice.lower() == "cancel":
        print("\nPDF splitting canceled. Thank you!")
        return

    try:
        pdf_index = int(pdf_choice) - 1
        if pdf_index < 0 or pdf_index >= len(pdf_files):
            print("Invalid PDF number.")
            return

        selected_pdf = pdf_files[pdf_index]
        input_pdf_path = os.path.join(folder_path, selected_pdf)

        pages_to_split = input("\nEnter page numbers or ranges (e.g., 1-3, 5, 7-9):\n")

        # Generate the output PDF path based on the input PDF file name
        output_pdf_name = f"Split_{os.path.splitext(selected_pdf)[0]}_{curr_time}.pdf"
        output_pdf_path = os.path.join(folder_path, output_pdf_name)

        split_pdf(input_pdf_path, output_pdf_path, pages_to_split)
        print(f"\nPDF pages split and saved to {output_pdf_path}")

    except ValueError:
        print("Invalid input.")

def mergePDF():
    curr_time = datetime.now().strftime("%Y%m%d%H%M%S")
    # Specify the directory containing the PDF files
    pdf_directory = input("Enter the directory containing the PDF files:\n")
    # Get the folder name
    folder_name = os.path.basename(pdf_directory)
    # Get a list of PDF files in the directory and sort them by filename
    pdf_files = [pdf_file for pdf_file in os.listdir(pdf_directory) if pdf_file.endswith('.pdf')]
    pdf_files.sort()

    if not pdf_files:
        print("\nNo PDF files found in the directory.")
        return

    print("\nPDF files found in the directory:\n")
    for pdf_file in pdf_files:
        print(pdf_file)

    confirmation = input("\nDo you want to merge these PDFs? (yes/no): ")
    if confirmation.lower() != "yes":
        print("\nPDF merging canceled. Thank you!")
        return

    # Create a PdfMerger object
    pdf_merger = PdfMerger()
    # Iterate through sorted PDF files and append them to the merger
    for pdf_file in pdf_files:
        pdf_path = os.path.join(pdf_directory, pdf_file)
        pdf_merger.append(pdf_path)
    
    # Specify the output PDF file path
    output_path = os.path.join(pdf_directory, folder_name+"_" + curr_time + ".pdf")
    # Write the merged PDF to the output file
    with open(output_path, 'wb') as output_file:
        pdf_merger.write(output_file)
    print("\nPDFs merged successfully. Thank you!")

def cleanExcel():
    # Create an empty list to store the input and output file names
    file_names = []

    # Ask the user for the file names and add them to the list
    while True:
        file_name_in = input("Enter excel file name (or press Enter to finish): ")
        if not file_name_in:
            break
        file_name_out = input("Enter output file name: ")
        file_names.append((file_name_in, file_name_out))

    # Get all file name in joined
    all_file_name = "\n".join(str(in_file) for in_file in file_names)

    print('\nYou have input these files: \n\n'+all_file_name)

    # Loop through the list of file names and clean each file
    for file_name_in, file_name_out in file_names:

        file_name_in_color = Back.BLUE + file_name_in + Style.RESET_ALL
        print("\nCleaning Excel File: ", file_name_in_color)

        # Read the Excel file
        try:
            df = pd.read_excel(file_name_in+'.xlsx')
        except FileNotFoundError:
            print("\n"+Back.RED+f"File '{file_name_in}.xlsx' not found. Skipping file..."+Style.RESET_ALL)
            continue
        
        # Remove formatting and objects
        wb = load_workbook(file_name_in+'.xlsx')
        ws = wb.active
        for row in ws.iter_rows():
            for cell in row:
                cell.value = cell.value

        # Write the cleaned data to a new Excel file
        df.to_excel(file_name_out+'.xlsx', index=False)
        print("Excel file " +file_name_in_color+ " cleaning done.")

    print("\nCleaning excel files have been completed. Thank you !")

def mergeExcel():
    # Create an empty list to store the input file names
    input_files = []

    # Ask the user to enter the input file names one by one, until they enter 'done'
    while True:
        filename = input("Enter excel file name (or press Enter to finish): ")
        if not filename:
            break
        input_files.append(filename)

    # Get the output file name from the user
    filename_output = input("\nEnter the output file name: ")

    # Get all file name in joined
    all_file_name = "\n".join(str(in_file) for in_file in input_files)

    print('\nYou have input these files: \n\n'+all_file_name)
    print("\nExcel file " +Back.BLUE + filename_output + Style.RESET_ALL+ " merging started, please wait....")

    # Create a new output workbook
    workbook_output = openpyxl.Workbook()

    # Get the active worksheet of the output workbook
    worksheet_output = workbook_output.active

    # Loop through each input file and copy its rows to the output worksheet
    for i, input_file in enumerate(input_files):
        # Load the input workbook into memory and select the first worksheet
        try:
            workbook = openpyxl.load_workbook(input_file + ".xlsx")
        except FileNotFoundError:
            print("\n"+Back.RED+f"File '{input_file}.xlsx' not found. Skipping file..."+Style.RESET_ALL+"\n")
            continue

        worksheet = workbook.worksheets[0]

        # If this is the first input file, append its entire first row (header row) to the output worksheet
        if i == 0:
            header_row = [cell.value for cell in worksheet[1]]
            worksheet_output.append(header_row)

        # Append all the rows from the input worksheet (except the first row, which contains headers)
        # to the output worksheet
        for row in worksheet.iter_rows(min_row=2):
            # Get the values of each cell in the row and add them to a list
            row_values = [cell.value for cell in row]
            # Append the list of row values to the output worksheet
            worksheet_output.append(row_values)

    # Save the output workbook to the specified filename
    workbook_output.save(filename_output + ".xlsx")

    print("Merging Excel Files have Completed. Thank You !")

def clean_and_merge():
    # Create an empty list to store the input and output file names
    file_names = []

    # Ask the user for the file names and add them to the list
    while True:
        file_name_in = input("\nEnter excel file name (or press Enter to finish): ")
        if not file_name_in:
            break
        file_name_out = input("Enter output file name: ")
        file_names.append((file_name_in, file_name_out))

    # Merged output file name
    merged_output = input("Merged File name: ")

    # Get all file name in joined
    all_file_name = "\n".join(str(in_file) for in_file in file_names)

    print('\nYou have input these files: \n\n'+all_file_name)
    print("\nNew File Name: " +Back.BLUE + merged_output + Style.RESET_ALL+"")

    # Create a new output workbook
    workbook_output = openpyxl.Workbook()

    # Get the active worksheet of the output workbook
    worksheet_output = workbook_output.active

    # Loop through the list of file names, clean each file, and copy its rows to the output worksheet
    for file_name_in, file_name_out in file_names:

        file_name_in_color = Back.BLUE + file_name_in + Style.RESET_ALL
        print("\nCleaning Excel File: ", file_name_in_color)

        # Read the Excel file
        try:
            df = pd.read_excel(file_name_in+'.xlsx')
        except FileNotFoundError:
            print("\n"+Back.RED+f"File '{file_name_in}.xlsx' not found. Skipping file..."+Style.RESET_ALL)
            continue

        # Remove formatting and objects
        wb = openpyxl.load_workbook(file_name_in+'.xlsx')
        ws = wb.active
        for row in ws.iter_rows():
            for cell in row:
                cell.value = cell.value

        # Write the cleaned data to a new Excel file
        df.to_excel(file_name_out+'_reformated.xlsx', index=False)

        print("Excel file " +file_name_in_color+ " cleaning done.")

        # Load the input workbook into memory and select the first worksheet
        workbook = openpyxl.load_workbook(file_name_out + "_reformated.xlsx")
        worksheet = workbook.worksheets[0]

        # If this is the first input file, append its entire first row (header row) to the output worksheet
        if file_names.index((file_name_in, file_name_out)) == 0:
            header_row = [cell.value for cell in worksheet[1]]
            worksheet_output.append(header_row)

        # Append all the rows from the input worksheet (except the first row, which contains headers)
        # to the output worksheet
        for row in worksheet.iter_rows(min_row=2):
            # Get the values of each cell in the row and add them to a list
            row_values = [cell.value for cell in row]
            # Append the list of row values to the output worksheet
            worksheet_output.append(row_values)

    # Save the output workbook to the specified filename
    print("\nExcel file " +Back.BLUE + merged_output + Style.RESET_ALL+ " merging started, please wait....")
    workbook_output.save(merged_output+".xlsx")

    print("\nCleaning and merging excel files have been completed. Thank you !")

"""
def filter_and_merge():
    # Create an empty list to store the input file names
    input_files = []

    # Ask the user to enter the input file names one by one, until they enter 'done'
    while True:
        filename = input("Enter excel file name (or press Enter to finish): ")
        if not filename:
            break
        input_files.append(filename)

    # Set the name of the column to filter and the keyword to look for
    column_to_filter = input("\nEnter filter column name: ")
    keyword_to_filter = input("Enter "+column_to_filter+" filter keyword: ")

    # Get all file name in joined
    all_file_name = "\n".join(str(in_file) for in_file in input_files)

    print('\nYou have input these files: \n\n'+all_file_name)
    print("\nFiltering and merging started, please wait....")

    # Initialize a list to store the filtered dataframes
    dfs = []

    # Loop through the Excel files and filter the desired column by the keyword
    for files in input_files:
        try:
            df = pd.read_excel(files+".xlsx", sheet_name=0)
        except FileNotFoundError:
            print("\n"+Back.RED+f"File '{files}.xlsx' not found. Skipping file..."+Style.RESET_ALL)
            continue

        df_filtered = df[df[column_to_filter] == keyword_to_filter]
        dfs.append(df_filtered)

    # Concatenate the filtered dataframes into a single dataframe
    merged_df = pd.concat(dfs, axis=0)

    # Save the merged dataframe to a new Excel file
    with pd.ExcelWriter(keyword_to_filter+'.xlsx') as writer:
        merged_df.to_excel(writer, index=False, sheet_name=keyword_to_filter)

    print('\nFiltered in '+Back.BLUE + column_to_filter + Style.RESET_ALL+' (column) by '+Back.BLUE + keyword_to_filter + Style.RESET_ALL+' (column keyword) and excel file merged done. Thank you !')
"""

def filter_and_merge_key():
    # Create an empty list to store the input file names
    input_files = []

    # Ask the user to enter the input file names one by one, until they enter 'done'
    while True:
        filename = input("Enter excel file name (or press Enter to finish): ")
        if not filename:
            break
        input_files.append(filename)

    # Set the name of the column to filter and the keywords to look for
    column_to_filter = input("\nEnter filter column name: ")
    keywords_to_filter = input("Enter " + column_to_filter + " filter keywords (separated by commas): ").split(",")

    # Get all file name in joined
    all_file_name = "\n".join(str(in_file) for in_file in input_files)

    print('\nYou have input these files: \n\n' + all_file_name)
    print("\nFiltering and merging started, please wait....")

    # Initialize a list to store the filtered dataframes
    dfs = []

    # Loop through the Excel files and filter the desired column by the keywords
    for file in input_files:
        try:
            df = pd.read_excel(file + ".xlsx", sheet_name=0)
        except FileNotFoundError:
            print("\n" + Back.RED + f"File '{file}.xlsx' not found. Skipping file..." + Style.RESET_ALL)
            continue
        
        if not keywords_to_filter:
            df_filtered = df[df[column_to_filter]==""]
        else:
            df_filtered = df[df[column_to_filter].isin(keywords_to_filter)]

        dfs.append(df_filtered)

    # Concatenate the filtered dataframes into a single dataframe
    merged_df = pd.concat(dfs, axis=0)

    # Save the merged dataframe to a new Excel file
    now = datetime.now()
    time_str = now.strftime("%Y%m%d_%H%M%S")
    output_file_name = "_".join(keywords_to_filter) + "_" + time_str + ".xlsx"
    with pd.ExcelWriter(output_file_name) as writer:
        merged_df.to_excel(writer, index=False, sheet_name="Filtered")

    print('\nFiltered in ' + Back.BLUE + column_to_filter + Style.RESET_ALL + ' (column) by ' +
          Back.BLUE + ", ".join(keywords_to_filter) + Style.RESET_ALL + ' (column keywords) and excel file merged done. Thank you !') 

def merge_sheets():
    # Create an empty list to store the input file names
    excel_file_names = []

    # Ask the user to enter the input file names one by one, until they enter 'done'
    while True:
        filename = input("Enter excel file name (or press Enter to finish): ")
        if not filename:
            break
        excel_file_names.append(filename)

    # Get all file name in joined
    all_file_name = "\n".join(str(in_file) for in_file in excel_file_names)

    print('\nYou have input these files: \n\n' + all_file_name)
    print("\nSheets merging started, please wait....")

    # Create a new sheet to merge all the sheets into.
    merged_sheet = pd.DataFrame()

    # Iterate over all the Excel files and merge their sheets into the new sheet.
    for excel_file in excel_file_names:
        sheets = pd.ExcelFile(excel_file+".xlsx").sheet_names
        for sheet in sheets:
            merged_sheet = pd.concat([merged_sheet, pd.read_excel(excel_file+".xlsx", sheet_name=sheet)], ignore_index=True)

        # Save each merged sheet to a new Excel file
        now = dt.datetime.now()
        time_str = now.strftime("%Y%m%d_%H%M%S")
        merged_sheet.to_excel(f"{excel_file}_{time_str}.xlsx", sheet_name="merged_sheet")
        merged_sheet = pd.DataFrame()  # Reset the merged_sheet DataFrame for the next file


    """
    # Save the Excel files with the current date.
    for i in enumerate(excel_file_names):
        # Save the merged dataframe to a new Excel file
        now = dt.datetime.now()
        time_str = now.strftime("%Y%m%d_%H%M%S")
        merged_sheet.to_excel(f"{i[1]}_{time_str}.xlsx", sheet_name="merged_sheet")
    """

    print('\nSheets of excel file '+Back.BLUE + ", ".join(excel_file_names) + Style.RESET_ALL+' merged done. Thank you !') 

# Call the initial function
try:
    initial()
except KeyboardInterrupt:
    print ("\n\n"+Back.RED+"Keyboard interrupted ! Exiting by the user. Thank you !"+Style.RESET_ALL+"\n")
    exit(1)